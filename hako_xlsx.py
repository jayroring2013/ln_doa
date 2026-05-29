#!/usr/bin/env python3
"""
LiDex — Novel Sync (licensed_books.sql → Supabase)

Parses a MariaDB dump of Vietnamese licensed light novels from Hako/Docln,
groups volumes by series_code, optionally enriches with AniList metadata,
and writes to: series, novel_meta, volumes, series_links

Data source: licensed_books table from Hako (docln) MariaDB database.
Each row = one published Vietnamese volume of a light novel.

Field mapping:
  licensed_books          →  LiDex
  ──────────────────────────────────────
  series_code (group)     →  series.title, series.title_vi, series.slug
  summary (first vol)     →  series.description_vi
  cover (first vol)       →  series.cover_url
  ── per-volume ──
  title                   →  volumes.title
  isbn                    →  volumes.isbn
  cover                   →  volumes.cover_url
  price                   →  volumes.price
  released_at             →  volumes.release_date
  translator              →  volumes.translator  (new column)
  page_count              →  volumes.page_count  (new column)
  retailers (JSON)        →  series_links (link_type='purchase')

Usage:
  python sync_novels.py                              # parse & import
  python sync_novels.py --sql=licensed_books.sql     # specify SQL file
  python sync_novels.py --anilist                    # also search AniList
  python sync_novels.py --dry-run                    # parse only, don't write
  python sync_novels.py --skip-existing              # skip series already in DB
"""

import sys, os, re, json, time, argparse
from urllib.parse import quote
from datetime import datetime, timezone, date
from collections import OrderedDict
from sync_common import (
    SUPABASE_URL, SUPABASE_KEY,
    _session, fetch_json, get_existing_ids, upsert_batch,
)

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass

# ── AniList GraphQL ─────────────────────────────────────────────────────

ANILIST_SEARCH_QUERY = """
query($search: String, $type: MediaType) {
  Media(search: $search, type: $type) {
    id
    idMal
    title { romaji english native }
    coverImage { extraLarge medium }
    bannerImage
    description(asHtml: false)
    genres
    status
    synonyms
    staff {
      edges(role: ORIGINAL_CREATOR) {
        node { name { full } }
      }
    }
  }
}
"""

ANILIST_STATUS_MAP = {
    'FINISHED':         'completed',
    'RELEASING':        'releasing',
    'NOT_YET_RELEASED': 'not_yet_released',
    'CANCELLED':        'cancelled',
    'HIATUS':           'hiatus',
}

# ── Vietnamese volume title parsing ────────────────────────────────────

VOL_NUM_PATTERN = re.compile(r'\s*[-\u2013\u2014]\s*T\u1eadp\s+(\d+)',
                            re.IGNORECASE)
SPECIAL_KEYWORDS = re.compile(
    r'\b(T\u1eadp \u0111\u1eb7c bi\u1ec7t|Special|Omnibus|'
    r'T\u1ed5ng h\u1ee3p|Artbook|Fanbook|Guide)\b',
    re.IGNORECASE,
)

# ── Known series_code → AniList search term mapping ────────────────────
KNOWN_ALIASES = {
    'dibiet':   'Mahouka Koukou no Rettousei',
    'cgvc':     'Chuunibyou demo Koi ga Shitai',
    'soigv':    'Seirei Tsukai no Blade Dance',
    'saop':     'Sword Art Online Progressive',
    'dal':      'Durarara',
    'cote':     'Classroom of the Elite',
    'aobuta':   'Seishun Buta Yarou',
    'bsd':      'Bungou Stray Dogs',
    'fz':       'Fate Zero',
    'gjbu':     'GJ-bu',
    'honzuki':  'Honzuki no Gekokujou',
    'jbf':      'Jaku-Chara Tomozaki-kun',
    'mvkt':     'Muv-Luv Alternative',
    '02lux':    'Fate Grand Order',
    'tsw':      'The Tatami Galaxy',
    'biblia':   'Biblia Koshodou no Jiken Techou',
    'spyroom':  'Spy x Family',
    'otomemob': 'Otomemosuki ni Ikiru',
    'hard':     'Hard Luck',
    'soft':     'Soft',
    'inobato':  'Imouto sae Ireba Ii',
    'sakurako': 'Sakurako-san no Ashimoto ni wa Shitai ga Umatteiru',
    'mahouiku': 'Mahou Shoujo Ikusei Keikaku',
    'gainga':   'Gangsta',
    'hyouka':   'Hyouka',
    'tenki':    'Tenki no Ko',
    'monogatari': 'Monogatari Series',
    'kumobasho': 'Kumo desu ga Nani ka',
    'grimgar':  'Hai to Gensou no Grimgar',
}


# ═══════════════════════════════════════════════════════════════════════
# 1. SQL PARSER — MariaDB dump → Python dicts (custom tokenizer)
# ═══════════════════════════════════════════════════════════════════════

def parse_licensed_books(sql_path):
    """
    Parse a MariaDB dump of the licensed_books table.
    Uses a custom character-by-character tokenizer to handle MariaDB's
    \\' escaping (which is incompatible with SQLite's executescript).
    Returns list of dicts, one per volume.
    """
    print(f'\n  Parsing SQL: {sql_path}')

    with open(sql_path, 'r', encoding='utf-8') as f:
        raw_sql = f.read()

    values_idx = re.search(
        r'INSERT\s+INTO\s+`?\w+`?\s*VALUES\s*',
        raw_sql, re.IGNORECASE)
    if not values_idx:
        print('  ERROR: Could not find INSERT VALUES in SQL file')
        return []

    values_text = raw_sql[values_idx.end():]
    values_text = values_text.strip()
    if values_text.endswith(';'):
        values_text = values_text[:-1].strip()

    FIELD_NAMES = [
        'id', 'user_id', 'title', 'slug', 'series_id', 'series_code',
        'translator', 'price', 'page_count', 'summary', 'isbn',
        'cover_type', 'cover', 'is_tba', 'retailers', 'released_at',
        'created_at', 'updated_at', 'preview', 'has_preview', 'view_count',
    ]
    NUM_FIELDS = len(FIELD_NAMES)

    rows = []
    i = 0
    text_len = len(values_text)

    while i < text_len:
        while i < text_len and values_text[i] in ' \n\r\t,':
            i += 1
        if i >= text_len or values_text[i] != '(':
            if i < text_len:
                i += 1
            continue

        i += 1
        fields = []

        while i < text_len:
            while i < text_len and values_text[i] in ' \n\r\t':
                i += 1
            if i >= text_len:
                break

            ch = values_text[i]

            if ch == ')':
                i += 1
                break
            elif ch == ',':
                i += 1
                continue
            elif ch == "'":
                i += 1
                field_chars = []
                while i < text_len:
                    c = values_text[i]
                    if c == '\\' and i + 1 < text_len and values_text[i+1] == "'":
                        field_chars.append("'")
                        i += 2
                    elif c == "'" and i + 1 < text_len and values_text[i+1] == "'":
                        field_chars.append("'")
                        i += 2
                    elif c == '\\' and i + 1 < text_len:
                        next_c = values_text[i + 1]
                        if next_c == 'n':
                            field_chars.append('\n')
                        elif next_c == 'r':
                            field_chars.append('\r')
                        elif next_c == 't':
                            field_chars.append('\t')
                        elif next_c == '\\':
                            field_chars.append('\\')
                        else:
                            field_chars.append(next_c)
                        i += 2
                    elif c == "'":
                        i += 1
                        break
                    else:
                        field_chars.append(c)
                        i += 1
                fields.append(''.join(field_chars))
            elif ch == 'N' and i + 3 < text_len and values_text[i:i+4] == 'NULL':
                fields.append(None)
                i += 4
            else:
                start = i
                while i < text_len and values_text[i] not in ',)':
                    i += 1
                token = values_text[start:i].strip()
                if token:
                    try:
                        if '.' in token:
                            fields.append(float(token))
                        else:
                            fields.append(int(token))
                    except ValueError:
                        fields.append(token)
                else:
                    fields.append(None)

        if len(fields) == NUM_FIELDS:
            row = dict(zip(FIELD_NAMES, fields))
            rows.append(row)
        elif fields and len(fields) > 5:
            print(f'  Warning: row {fields[0]} has {len(fields)} fields '
                  f'(expected {NUM_FIELDS}), skipping')

    print(f'  Parsed {len(rows)} volume rows from SQL dump')
    return rows


# ═══════════════════════════════════════════════════════════════════════
# 2. GROUP VOLUMES BY SERIES
# ═══════════════════════════════════════════════════════════════════════

def extract_volume_number(title):
    if not title:
        return None, title, False

    is_special = bool(SPECIAL_KEYWORDS.search(title))

    match = VOL_NUM_PATTERN.search(title)
    if match:
        vol_num = int(match.group(1))
        series_name = title[:match.start()].strip().rstrip('-').strip()
        return vol_num, series_name, False

    return None, title, is_special


def group_volumes_by_series(rows):
    series = OrderedDict()

    for row in rows:
        code = row['series_code']
        vol_num, series_name, is_special = extract_volume_number(row['title'])

        if code not in series:
            series[code] = {
                'volumes': [],
                'series_name': series_name,
                'first_volume': row,
            }

        series[code]['volumes'].append({
            **row,
            '_vol_num': vol_num,
            '_is_special': is_special or (vol_num is None),
            '_series_name': series_name,
        })

    return series


# ═══════════════════════════════════════════════════════════════════════
# 3. RETAILER LINK PARSER
# ═══════════════════════════════════════════════════════════════════════

RETAILER_LABELS = {
    'tiki':    'Tiki',
    'fahasa':  'Fahasa',
    'shopee':  'Shopee',
    'pibook':  'PiBook',
    'hikaru':  'Hikaru Books',
    'amazon':  'Amazon',
}


def parse_retailer_links(retailers_json):
    if not retailers_json:
        return []
    try:
        data = json.loads(retailers_json)
    except (json.JSONDecodeError, TypeError):
        return []

    links = []
    if isinstance(data, dict):
        for retailer, urls in data.items():
            if not isinstance(urls, list):
                continue
            for url in urls:
                if url and isinstance(url, str) and url.startswith('http'):
                    label = RETAILER_LABELS.get(retailer.lower(),
                                                retailer.title())
                    links.append({'label': label, 'url': url})
    elif isinstance(data, list):
        for url in data:
            if url and isinstance(url, str) and url.startswith('http'):
                links.append({'label': 'Purchase', 'url': url})

    return links


# ═══════════════════════════════════════════════════════════════════════
# 4. ANILIST ENRICHMENT (optional)
# ═══════════════════════════════════════════════════════════════════════

def search_anilist(search_term):
    if not search_term:
        return None
    try:
        data = fetch_json(
            'https://graphql.anilist.co',
            method='POST',
            body={
                'query': ANILIST_SEARCH_QUERY,
                'variables': {'search': search_term, 'type': 'NOVEL'},
            },
            timeout=15,
        )
        if not data or not data.get('data'):
            return None
        return data['data'].get('Media')
    except Exception as e:
        print(f'\n  AniList error for "{search_term}": {e}')
        return None


def enrich_series_from_anilist(series_data):
    total = len(series_data)
    print(f'\n  AniList enrichment - searching {total} series...')
    print(f'  (Rate limit: ~1s per request, ~{total}s estimated)\n')

    stats = {'found': 0, 'not_found': 0, 'total': total}

    for idx, (code, info) in enumerate(series_data.items()):
        search_term = KNOWN_ALIASES.get(code, code)
        vn_name = info['series_name']

        print(f'  [{idx+1}/{total}] {code} -> "{search_term}"...',
              end=' ', flush=True)

        media = search_anilist(search_term)

        if not media and search_term != vn_name:
            print('(retry with VN)...', end=' ', flush=True)
            media = search_anilist(vn_name)

        if media:
            stats['found'] += 1
            title = media.get('title', {})
            staff = media.get('staff', {})
            edges = staff.get('edges') or [{}]
            author = (edges[0].get('node') or {}).get('name', {}).get('full')

            info['anilist'] = {
                'anilist_id':   media.get('id'),
                'mal_id':       media.get('idMal'),
                'title':        title.get('english') or title.get('romaji', ''),
                'title_native': title.get('native'),
                'title_english': title.get('english'),
                'cover_url':    (media.get('coverImage') or {}).get('extraLarge'),
                'banner_url':   media.get('bannerImage'),
                'description':  (media.get('description') or '')[:2000],
                'genres':       media.get('genres', []),
                'status':       ANILIST_STATUS_MAP.get(media.get('status'),
                                                    'unknown'),
                'author':       author,
            }
            print(f'OK (AL:{media["id"]})')
        else:
            stats['not_found'] += 1
            print('NOT FOUND')

        time.sleep(1.0)

    print(f'\n  AniList done: {stats["found"]} found, '
          f'{stats["not_found"]} not found')
    return stats


# ═══════════════════════════════════════════════════════════════════════
# 5. BUILD SUPABASE ROWS
# ═══════════════════════════════════════════════════════════════════════

def _strip_html(html_text):
    if not html_text:
        return ''
    text = re.sub(r'<[^>]+>', '', html_text)
    text = text.replace('&amp;', '&').replace('&lt;', '<')
    text = text.replace('&gt;', '>').replace('&quot;', '"')
    text = text.replace('&#39;', "'")
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _slugify(text):
    if not text:
        return None
    s = text.lower().strip()
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s-]+', '-', s)
    s = s.strip('-')
    return s if len(s) >= 2 else None


def _blank_to_none(value):
    if value == '':
        return None
    return value


def _excel_datetime(value, date_only=False):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat() if date_only else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None

    if date_only:
        return text[:10]
    if '+00:00' in text or text.endswith('Z'):
        return text
    if len(text) == 19 and text[10] == ' ':
        return text + '+00:00'
    return text


def _excel_array(value):
    if value in (None, '', '{}'):
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        text = value.strip()
        if text == '{}':
            return []
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return parsed
        except json.JSONDecodeError:
            pass
        if text.startswith('{') and text.endswith('}'):
            inner = text[1:-1].strip()
            return [x.strip().strip('"') for x in inner.split(',') if x.strip()]
    return []


def _normalize_series_status(value):
    status = (value or '').strip().lower()
    return {
        'completed': 'completed',
        'ongoing': 'ongoing',
        'dropped': 'cancelled',
        'stalled': 'hiatus',
        'caught_up': 'ongoing',
        'unknown': 'ongoing',
    }.get(status, 'ongoing')


def _load_xlsx_sheet_rows(xlsx_path, sheet_name):
    if load_workbook is None:
        raise RuntimeError('openpyxl is required for --xlsx imports')

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f'Workbook missing required sheet: {sheet_name}')

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else '' for h in next(rows)]
    output = []

    for row in rows:
        item = {}
        has_value = False
        for key, value in zip(headers, row):
            if not key:
                continue
            value = _blank_to_none(value)
            if value is not None:
                has_value = True
            item[key] = value
        if has_value:
            output.append(item)

    return output


def load_xlsx_import_rows(xlsx_path):
    print(f'\n  Reading Excel workbook: {xlsx_path}')
    series_rows = _load_xlsx_sheet_rows(xlsx_path, 'series')
    volume_rows = _load_xlsx_sheet_rows(xlsx_path, 'volumes')

    print(f'  Excel series rows:  {len(series_rows)}')
    print(f'  Excel volume rows:  {len(volume_rows)}')
    return series_rows, volume_rows


def build_series_rows_from_xlsx(raw_rows):
    rows = []
    workbook_id_to_slug = {}

    for raw in raw_rows:
        workbook_id = raw.get('id')
        slug = raw.get('slug') or _slugify(raw.get('title'))
        if not slug:
            continue
        if workbook_id is not None:
            workbook_id_to_slug[workbook_id] = slug

        rows.append({
            'item_type':      raw.get('item_type') or 'novel',
            'title':          raw.get('title'),
            'title_vi':       raw.get('title_vi') or raw.get('title'),
            'title_native':   raw.get('title_native'),
            'title_english':  raw.get('title_english'),
            'slug':           slug,
            'cover_url':      raw.get('cover_url'),
            'banner_url':     raw.get('banner_url'),
            'description':    raw.get('description'),
            'description_vi': raw.get('description_vi'),
            'status':         _normalize_series_status(raw.get('status')),
            'genres':         _excel_array(raw.get('genres')),
            'tags':           _excel_array(raw.get('tags')),
            'source':         raw.get('source') or 'LIGHT_NOVEL',
            'author':         raw.get('author'),
            'studio':         raw.get('studio'),
            'publisher_id':   raw.get('publisher_id'),
            'anilist_id':     raw.get('anilist_id'),
            'mangadex_id':    raw.get('mangadex_id'),
            'mal_id':         raw.get('mal_id'),
            'created_at':     _excel_datetime(raw.get('created_at')),
            'updated_at':     _excel_datetime(raw.get('updated_at')),
        })

    return rows, workbook_id_to_slug


def build_volume_rows_from_xlsx(raw_rows, workbook_id_to_db_id):
    rows = []

    for raw in raw_rows:
        workbook_series_id = raw.get('series_id')
        sid = workbook_id_to_db_id.get(workbook_series_id)
        if not sid:
            continue

        rows.append({
            'series_id':      sid,
            'volume_number':  raw.get('volume_number'),
            'title':          raw.get('title'),
            'isbn':           raw.get('isbn'),
            'cover_url':      raw.get('cover_url'),
            'release_date':   _excel_datetime(raw.get('release_date'), date_only=True),
            'price':          raw.get('price'),
            'currency':       raw.get('currency') or 'VND',
            'is_special':     bool(raw.get('is_special')),
            'is_digital':     bool(raw.get('is_digital')),
            'created_at':     _excel_datetime(raw.get('created_at')),
            'page_count':     raw.get('page_count'),
            'translator':     raw.get('translator'),
        })

    return rows


def build_novel_meta_rows_from_xlsx(series_rows, volume_rows, workbook_id_to_db_id):
    now = datetime.now(timezone.utc).isoformat()
    status_by_workbook_id = {row.get('id'): row.get('status') for row in series_rows}
    max_volume_by_db_id = {}

    for raw in volume_rows:
        sid = workbook_id_to_db_id.get(raw.get('series_id'))
        if not sid or raw.get('is_special'):
            continue
        vol_num = raw.get('volume_number')
        if vol_num is None:
            continue
        max_volume_by_db_id[sid] = max(max_volume_by_db_id.get(sid, 0), int(vol_num))

    rows = []
    for workbook_id, sid in workbook_id_to_db_id.items():
        status = (status_by_workbook_id.get(workbook_id) or '').lower()
        rows.append({
            'series_id':    sid,
            'volume_count': max_volume_by_db_id.get(sid, 0),
            'is_completed': status == 'completed',
            'updated_at':   now,
        })

    return rows


def build_series_rows(series_data):
    now = datetime.now(timezone.utc).isoformat()
    rows = []

    for code, info in series_data.items():
        al = info.get('anilist', {})
        first = info['first_volume']
        vn_name = info['series_name']

        title = al.get('title') or vn_name

        if al.get('title_english'):
            slug = _slugify(al['title_english'])
        elif al.get('title') and any(c.isascii() for c in al['title']):
            slug = _slugify(al['title'])
        else:
            slug = code

        description = al.get('description')
        description_vi = _strip_html(first.get('summary', ''))[:2000]

        status = al.get('status')
        if not status or status == 'unknown':
            max_vol = max(
                (v['_vol_num'] for v in info['volumes'] if v['_vol_num']),
                default=0,
            )
            status = 'completed' if max_vol > 0 else 'unknown'

        rows.append({
            'item_type':      'novel',
            'title':          title,
            'title_vi':       vn_name,
            'title_native':   al.get('title_native'),
            'title_english':  al.get('title_english'),
            'slug':           slug,
            'cover_url':      al.get('cover_url') or first.get('cover'),
            'banner_url':     al.get('banner_url'),
            'description':    description,
            'description_vi': description_vi,
            'status':         status,
            'genres':         al.get('genres', []),
            'tags':           [],
            'source':         'LIGHT_NOVEL',
            'author':         al.get('author'),
            'anilist_id':     al.get('anilist_id'),
            'mal_id':         al.get('mal_id'),
            'updated_at':     now,
        })

    return rows


def build_novel_meta_rows(series_data, series_id_map):
    now = datetime.now(timezone.utc).isoformat()
    rows = []

    for code, info in series_data.items():
        sid = series_id_map.get(code)
        if not sid:
            continue

        regular_vols = [v for v in info['volumes']
                        if v['_vol_num'] is not None and not v['_is_special']]
        max_vol = max((v['_vol_num'] for v in regular_vols), default=0)

        has_tba = any(v.get('is_tba') for v in info['volumes'])
        is_completed = (not has_tba and max_vol > 0)

        rows.append({
            'series_id':    sid,
            'volume_count': max_vol,
            'is_completed': is_completed,
            'updated_at':   now,
        })

    return rows


def build_volume_rows(series_data, series_id_map):
    now = datetime.now(timezone.utc).isoformat()
    rows = []

    for code, info in series_data.items():
        sid = series_id_map.get(code)
        if not sid:
            continue

        for vol in info['volumes']:
            release_date = None
            if vol.get('released_at'):
                try:
                    dt = datetime.strptime(vol['released_at'], '%Y-%m-%d %H:%M:%S')
                    release_date = dt.strftime('%Y-%m-%d')
                except (ValueError, TypeError):
                    pass

            rows.append({
                'series_id':     sid,
                'volume_number':  vol['_vol_num'],
                'title':          vol.get('title', ''),
                'isbn':           vol.get('isbn') or None,
                'cover_url':      vol.get('cover') or None,
                'release_date':   release_date,
                'price':          vol.get('price'),
                'currency':       'VND',
                'is_special':     vol['_is_special'],
                'is_digital':     False,
                'translator':     vol.get('translator') or None,
                'page_count':     vol.get('page_count') or None,
                'created_at':     now,
            })

    return rows


def build_series_link_rows(series_data, series_id_map):
    now = datetime.now(timezone.utc).isoformat()
    rows = []

    for code, info in series_data.items():
        sid = series_id_map.get(code)
        if not sid:
            continue

        seen_urls = set()
        link_list = []

        for vol in info['volumes']:
            links = parse_retailer_links(vol.get('retailers', ''))
            for link in links:
                url = link['url']
                if url not in seen_urls:
                    seen_urls.add(url)
                    link_list.append(link)

        for i, link in enumerate(link_list):
            rows.append({
                'series_id':  sid,
                'link_type':  'purchase',
                'label':      link['label'],
                'url':        link['url'],
                'is_active':  True,
                'sort_order': i + 1,
                'created_at': now,
            })

    return rows


# ═══════════════════════════════════════════════════════════════════════
# 6. SUPABASE FLUSH HELPERS
# ═══════════════════════════════════════════════════════════════════════

def _sb_headers():
    return {
        'apikey':        SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type':  'application/json',
        'Prefer':        'resolution=merge-duplicates,return=minimal',
    }


def _flush(rows, table, conflict, label=None):
    if not rows:
        return 0
    label = label or table
    url = f'{SUPABASE_URL}/rest/v1/{table}?on_conflict={conflict}'
    h = _sb_headers()
    total = 0

    batch_size = 50 if table == 'series' else 100
    for i in range(0, len(rows), batch_size):
        chunk = rows[i:i + batch_size]
        try:
            r = _session.post(url, json=chunk, headers=h, timeout=60)
            if r.status_code in (200, 201, 204):
                total += len(chunk)
            else:
                print(f'  {label} [{r.status_code}]: {r.text[:300]}')
        except Exception as e:
            print(f'  {label} upsert error: {e}')
        time.sleep(0.3)

    return total


def _resolve_series_by_slug(slugs):
    if not slugs:
        return {}
    h = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'}
    id_map = {}
    unique = list(set(slugs))
    for i in range(0, len(unique), 25):
        batch = unique[i:i + 25]
        slugs_param = ','.join(quote(str(slug), safe='') for slug in batch)
        try:
            r = _session.get(
                f'{SUPABASE_URL}/rest/v1/series?select=id,slug'
                f'&slug=in.({slugs_param})',
                headers=h, timeout=30,
            )
            if r.status_code == 200 and isinstance(r.json(), list):
                for row in r.json():
                    if isinstance(row, dict) and row.get('slug'):
                        id_map[row['slug']] = row['id']
            else:
                print(f'  series lookup [{r.status_code}]: {r.text[:200]}')
        except Exception as e:
            print(f'  series lookup error: {e}')
    return id_map


def _resolve_series_by_anilist(al_ids):
    if not al_ids:
        return {}
    h = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'}
    id_map = {}
    unique = list(set(al_ids))
    for i in range(0, len(unique), 200):
        batch = unique[i:i + 200]
        ids_param = ','.join(str(x) for x in batch)
        try:
            r = _session.get(
                f'{SUPABASE_URL}/rest/v1/series?select=id,anilist_id'
                f'&anilist_id=in.({ids_param})',
                headers=h, timeout=30,
            )
            if r.status_code == 200 and isinstance(r.json(), list):
                for row in r.json():
                    if isinstance(row, dict) and row.get('anilist_id'):
                        id_map[row['anilist_id']] = row['id']
        except Exception as e:
            print(f'  anilist lookup error: {e}')
    return id_map


# ═══════════════════════════════════════════════════════════════════════
# 7. SUMMARY
# ═══════════════════════════════════════════════════════════════════════

def _fetch_novel_series_lookup():
    h = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'}
    lookup = {'by_slug': {}, 'by_title_vi': {}, 'by_title': {}}
    offset = 0
    limit = 1000

    while True:
        try:
            r = _session.get(
                f'{SUPABASE_URL}/rest/v1/series'
                f'?select=id,slug,title,title_vi'
                f'&item_type=eq.novel'
                f'&limit={limit}&offset={offset}',
                headers=h,
                timeout=30,
            )
            if r.status_code != 200:
                print(f'  novel series lookup [{r.status_code}]: {r.text[:200]}')
                break
            rows = r.json()
            if not rows:
                break
            for row in rows:
                if row.get('slug'):
                    lookup['by_slug'][row['slug']] = row['id']
                if row.get('title_vi'):
                    lookup['by_title_vi'][row['title_vi']] = row['id']
                if row.get('title'):
                    lookup['by_title'][row['title']] = row['id']
            if len(rows) < limit:
                break
            offset += limit
        except Exception as e:
            print(f'  novel series lookup error: {e}')
            break

    return lookup


def print_summary(series_data):
    total_vols = sum(len(s['volumes']) for s in series_data.values())
    total_series = len(series_data)

    vol_ranges = {'1 vol': 0, '2-5': 0, '6-10': 0, '11-20': 0, '20+': 0}
    for code, info in series_data.items():
        n = len(info['volumes'])
        if n == 1: vol_ranges['1 vol'] += 1
        elif n <= 5: vol_ranges['2-5'] += 1
        elif n <= 10: vol_ranges['6-10'] += 1
        elif n <= 20: vol_ranges['11-20'] += 1
        else: vol_ranges['20+'] += 1

    total_links = 0
    for code, info in series_data.items():
        for vol in info['volumes']:
            total_links += len(parse_retailer_links(vol.get('retailers', '')))

    print(f'\n  {"="*52}')
    print(f'  Data Summary')
    print(f'  {"="*52}')
    print(f'  Total series:         {total_series}')
    print(f'  Total volumes:        {total_vols}')
    print(f'  Total retailer links: {total_links}')
    print(f'  {"-"*52}')
    print(f'  Distribution by volume count:')
    for label, count in vol_ranges.items():
        if count > 0:
            print(f'    {label:>8}: {count:>4} series')
    print(f'  {"-"*52}')
    print(f'  Top 15 series by volume count:')
    top15 = sorted(series_data.items(),
                   key=lambda x: len(x[1]['volumes']), reverse=True)[:15]
    for code, info in top15:
        n = len(info['volumes'])
        print(f'    {code:<16} {n:>3} vols - {info["series_name"][:40]}')
    print(f'  {"="*52}')


# ═══════════════════════════════════════════════════════════════════════
# 8. MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════════════

def run_xlsx_sync(xlsx_path, dry_run=False, skip_existing=False):
    start_time = time.time()

    print('=' * 60)
    print('  LiDex - Novel Sync (Excel workbook -> Supabase)')
    print(f'  {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print(f'  Excel file: {xlsx_path}')
    if dry_run:
        print('  Mode: DRY RUN (no writes)')
    if skip_existing:
        print('  Skip existing: ON')
    print('=' * 60)

    raw_series, raw_volumes = load_xlsx_import_rows(xlsx_path)
    series_rows, workbook_id_to_slug = build_series_rows_from_xlsx(raw_series)

    if dry_run:
        print(f'\n  === DRY RUN ===')
        print(f'  series:     {len(series_rows)} rows')
        print(f'  volumes:    {len(raw_volumes)} rows')
        print(f'  novel_meta: {len(series_rows)} rows')
        print(f'\n  Sample series rows (first 5):')
        for row in series_rows[:5]:
            print(f'    title={row["title"][:50]:<50} slug={row["slug"]}')
        elapsed = time.time() - start_time
        print(f'\n  Done in {elapsed:.1f}s - dry run complete')
        return

    existing_slugs = {}
    if skip_existing:
        print(f'\n  Checking existing series...')
        slugs = [row['slug'] for row in series_rows if row.get('slug')]
        existing_slugs = _resolve_series_by_slug(slugs)
        print(f'  Found {len(existing_slugs)} existing series by slug')

    new_series = [
        row for row in series_rows
        if not skip_existing or row.get('slug') not in existing_slugs
    ]

    if new_series:
        print(f'\n  series ({len(new_series)} new rows)...')
        _flush(new_series, 'series', 'slug', label='series')
    else:
        print(f'\n  series: 0 new rows (all already exist)')

    print(f'\n  Resolving series IDs...')
    all_slugs = [row['slug'] for row in series_rows if row.get('slug')]
    slug_to_id = _resolve_series_by_slug(all_slugs)
    novel_lookup = _fetch_novel_series_lookup()
    raw_series_by_workbook_id = {row.get('id'): row for row in raw_series}
    workbook_id_to_db_id = {
        workbook_id: (
            slug_to_id.get(slug)
            or novel_lookup['by_slug'].get(slug)
            or novel_lookup['by_title_vi'].get(
                raw_series_by_workbook_id.get(workbook_id, {}).get('title_vi'))
            or novel_lookup['by_title'].get(
                raw_series_by_workbook_id.get(workbook_id, {}).get('title'))
        )
        for workbook_id, slug in workbook_id_to_slug.items()
    }
    workbook_id_to_db_id = {
        workbook_id: sid
        for workbook_id, sid in workbook_id_to_db_id.items()
        if sid
    }

    unresolved = [
        workbook_id for workbook_id in workbook_id_to_slug
        if workbook_id not in workbook_id_to_db_id
    ]
    if unresolved:
        print(f'  Warning: {len(unresolved)} workbook series IDs unresolved')
        print(f'  First unresolved IDs: {unresolved[:10]}')

    meta_rows = build_novel_meta_rows_from_xlsx(
        raw_series, raw_volumes, workbook_id_to_db_id)
    if meta_rows:
        print(f'\n  novel_meta ({len(meta_rows)} rows)...')
        _flush(meta_rows, 'novel_meta', 'series_id')

    vol_rows = build_volume_rows_from_xlsx(raw_volumes, workbook_id_to_db_id)
    if vol_rows:
        print(f'\n  volumes ({len(vol_rows)} rows)...')
        _flush(vol_rows, 'volumes', 'series_id,volume_number')

    elapsed = time.time() - start_time
    print(f'\n  {"="*52}')
    print(f'  Excel sync complete!')
    print(f'  Series:  {len(series_rows)} total, {len(new_series)} new')
    print(f'  Volumes: {len(vol_rows)}')
    print(f'  Time:    {elapsed / 60:.1f} minutes')
    print(f'  {"="*52}')


def run_sync(sql_path, do_anilist=False, dry_run=False, skip_existing=False):
    start_time = time.time()

    print('=' * 60)
    print('  LiDex - Novel Sync (licensed_books.sql -> Supabase)')
    print(f'  {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print(f'  SQL file: {sql_path}')
    if do_anilist:
        print('  AniList enrichment: ON')
    if dry_run:
        print('  Mode: DRY RUN (no writes)')
    if skip_existing:
        print('  Skip existing: ON')
    print('=' * 60)

    # ── Step 1: Parse SQL ──
    rows = parse_licensed_books(sql_path)
    if not rows:
        print('\n  No data found. Exiting.')
        return

    # ── Step 2: Group by series ──
    series_data = group_volumes_by_series(rows)
    print_summary(series_data)

    # ── Step 3: AniList enrichment (optional) ──
    if do_anilist:
        enrich_series_from_anilist(series_data)

    # ── Step 4: Build rows ──
    print(f'\n  Building Supabase rows...')
    series_rows = build_series_rows(series_data)

    if dry_run:
        print(f'\n  === DRY RUN ===')
        print(f'  series:        {len(series_rows)} rows')
        total_vols = sum(len(s['volumes']) for s in series_data.values())
        print(f'  novel_meta:    {len(series_data)} rows')
        print(f'  volumes:       {total_vols} rows')

        print(f'\n  Sample series rows (first 5):')
        for row in series_rows[:5]:
            print(f'    title={row["title"][:50]:<50} '
                  f'slug={row["slug"]:<30} '
                  f'al={row.get("anilist_id")}')

        if do_anilist:
            mapping_path = os.path.join(
                os.path.dirname(sql_path), 'anilist_mapping.json')
            mapping = {}
            for code, info in series_data.items():
                al = info.get('anilist', {})
                mapping[code] = {
                    'vn_name': info['series_name'],
                    'volumes': len(info['volumes']),
                    **al,
                }
            with open(mapping_path, 'w', encoding='utf-8') as f:
                json.dump(mapping, f, ensure_ascii=False, indent=2)
            print(f'\n  AniList mapping saved to: {mapping_path}')

        elapsed = time.time() - start_time
        print(f'\n  Done in {elapsed:.1f}s - dry run complete')
        return

    # ── Step 5: Flush to Supabase ──
    print(f'\n  Flushing to Supabase...')

    existing_slugs = {}
    existing_al = {}
    if skip_existing:
        print(f'  Checking existing series...')
        slugs = [row['slug'] for row in series_rows if row.get('slug')]
        existing_slugs = _resolve_series_by_slug(slugs)
        al_ids = [row['anilist_id'] for row in series_rows
                  if row.get('anilist_id')]
        existing_al = _resolve_series_by_anilist(al_ids)
        print(f'  Found {len(existing_slugs)} by slug, '
              f'{len(existing_al)} by anilist_id')

    # Filter new series
    new_series = []
    for row in series_rows:
        skip = False
        if skip_existing:
            if row.get('slug') and row['slug'] in existing_slugs:
                skip = True
            if row.get('anilist_id') and row['anilist_id'] in existing_al:
                skip = True
        if not skip:
            new_series.append(row)

    if new_series:
        print(f'\n  series ({len(new_series)} new rows)...')
        _flush(new_series, 'series', 'slug', label='series')
    else:
        print(f'\n  series: 0 new rows (all already exist)')

    # ── Resolve series IDs ──
    print(f'\n  Resolving series IDs...')
    all_slugs = [row['slug'] for row in series_rows if row.get('slug')]
    slug_to_id = _resolve_series_by_slug(all_slugs)

    code_to_slug = {}
    for row in series_rows:
        for code in series_data:
            if series_data[code]['series_name'] == row.get('title_vi'):
                code_to_slug[code] = row['slug']
                break

    series_id_map = {}
    for code, slug in code_to_slug.items():
        sid = slug_to_id.get(slug)
        if sid:
            series_id_map[code] = sid

    unresolved = [c for c in series_data if c not in series_id_map]
    if unresolved:
        print(f'  Warning: {len(unresolved)} series unresolved:')
        for c in unresolved[:10]:
            print(f'    {c}: {series_data[c]["series_name"][:40]}')

    # ── Flush novel_meta ──
    meta_rows = build_novel_meta_rows(series_data, series_id_map)
    if meta_rows:
        print(f'\n  novel_meta ({len(meta_rows)} rows)...')
        _flush(meta_rows, 'novel_meta', 'series_id')

    # ── Flush volumes ──
    vol_rows = build_volume_rows(series_data, series_id_map)
    if vol_rows:
        print(f'\n  volumes ({len(vol_rows)} rows)...')
        _flush(vol_rows, 'volumes', 'series_id,volume_number')

    # ── Flush series_links ──
    link_rows = build_series_link_rows(series_data, series_id_map)
    if link_rows:
        print(f'\n  series_links ({len(link_rows)} rows)...')
        _flush(link_rows, 'series_links', 'series_id,url')

    # ── Done ──
    elapsed = time.time() - start_time
    print(f'\n  {"="*52}')
    print(f'  Sync complete!')
    print(f'  Series:  {len(series_rows)} total, {len(new_series)} new')
    print(f'  Volumes: {len(vol_rows)}')
    print(f'  Links:   {len(link_rows)}')
    print(f'  Time:    {elapsed / 60:.1f} minutes')
    print(f'  {"="*52}')


# ═══════════════════════════════════════════════════════════════════════
# 9. CLI
# ═══════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='LiDex Novel Sync - Import SQL or Excel data to Supabase')
    parser.add_argument('--sql', type=str, default='licensed_books.sql',
                        help='Path to licensed_books.sql MariaDB dump')
    parser.add_argument('--xlsx', type=str, default=r'D:\series_table_import_v5.xlsx',
                        help='Path to series_table_import_v5.xlsx workbook')
    parser.add_argument('--anilist', action='store_true',
                        help='Search AniList for metadata enrichment')
    parser.add_argument('--dry-run', action='store_true',
                        help='Parse and summarize without writing')
    parser.add_argument('--skip-existing', action='store_true',
                        help='Skip series that already exist in Supabase')
    args = parser.parse_args()

    if args.xlsx:
        xlsx_path = args.xlsx
        if not os.path.isabs(xlsx_path):
            xlsx_path = os.path.join(os.path.dirname(__file__), xlsx_path)

        if not os.path.exists(xlsx_path):
            print(f'ERROR: Excel file not found: {xlsx_path}')
            sys.exit(1)

        run_xlsx_sync(
            xlsx_path=xlsx_path,
            dry_run=args.dry_run,
            skip_existing=args.skip_existing,
        )
        sys.exit(0)

    sql_path = args.sql
    if not os.path.isabs(sql_path):
        sql_path = os.path.join(os.path.dirname(__file__), sql_path)

    if not os.path.exists(sql_path):
        print(f'ERROR: SQL file not found: {sql_path}')
        sys.exit(1)

    run_sync(
        sql_path=sql_path,
        do_anilist=args.anilist,
        dry_run=args.dry_run,
        skip_existing=args.skip_existing,
    )
