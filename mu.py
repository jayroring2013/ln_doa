import re
import json
import time
import random
import requests
from pathlib import Path
from urllib.parse import urljoin, urlparse, parse_qs, unquote
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

BASE = "https://www.mangaupdates.com"
START_URL = "https://www.mangaupdates.com/series.html?page={page}"
OUT_XLSX = "mangaupdates_manga_5000.xlsx"
SEEN_FILE = "seen_urls.json"
TARGET = 10

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; MangaMetadataTest/1.0)",
    "Accept-Language": "en-US,en;q=0.9",
}


def clean(text):
    return re.sub(r"\s+", " ", text or "").strip()


def clean_title(title):
    return re.sub(r"\s*-\s*MangaUpdates\s*$", "", title or "").strip()


def image_url(src):
    if not src:
        return None
    if "/_next/image" in src and "url=" in src:
        qs = parse_qs(urlparse(src).query)
        real = qs.get("url", [None])[0]
        if real:
            return urljoin(BASE, unquote(real))
    return urljoin(BASE, src)


def text_after_label(soup, label):
    node = soup.find(string=re.compile(rf"^\s*{re.escape(label)}\s*$", re.I))
    if not node:
        return None

    container = node.find_parent()
    for _ in range(6):
        sib = container.find_next_sibling() if container else None
        if sib:
            value = clean(sib.get_text(" ", strip=True))
            return value or None
        container = container.parent if container else None

    return None


def links_after_label(soup, label, href_contains=None, exclude_text=None):
    exclude_text = exclude_text or []

    node = soup.find(string=re.compile(rf"^\s*{re.escape(label)}\s*$", re.I))
    if not node:
        return []

    container = node.find_parent()
    for _ in range(6):
        sib = container.find_next_sibling() if container else None
        if sib:
            out = []
            for a in sib.find_all("a", href=True):
                txt = clean(a.get_text(" ", strip=True))
                href = a["href"]

                if not txt:
                    continue
                if href_contains and href_contains not in href:
                    continue
                if any(bad.lower() in txt.lower() for bad in exclude_text):
                    continue

                out.append(txt)

            return list(dict.fromkeys(out))

        container = container.parent if container else None

    return []


def parse_status(status_text):
    result = {"volumes": None, "chapters": None, "status": None, "raw": status_text}

    if not status_text:
        return result

    vol = re.search(r"(\d+)\s+Volumes?", status_text, re.I)
    chap = re.search(r"(\d+)\s+Chapters?", status_text, re.I)
    stat = re.search(r"\((.*?)\)", status_text)

    if vol:
        result["volumes"] = int(vol.group(1))
    if chap:
        result["chapters"] = int(chap.group(1))
    if stat:
        result["status"] = clean(stat.group(1))

    return result


def find_cover(soup):
    bad_words = ["mascot", "logo", "header", "baka-updates", "avatar", "icon", ".svg"]

    attrs = ["src", "data-src", "data-image"]

    for img in soup.find_all("img"):
        for attr in attrs:
            src = img.get(attr)
            if not src:
                continue

            full = image_url(src)
            lower = full.lower()

            if any(bad in lower for bad in bad_words):
                continue

            if any(ext in lower for ext in [".jpg", ".jpeg", ".png", ".webp"]):
                return full

    html = str(soup)

    patterns = [
        r'https://[^"]+\.(?:jpg|jpeg|png|webp)',
        r'https:\\/\\/[^"]+\.(?:jpg|jpeg|png|webp)',
    ]

    for pattern in patterns:
        for m in re.findall(pattern, html, re.I):
            m = m.replace("\\/", "/")
            lower = m.lower()

            if any(bad in lower for bad in bad_words):
                continue

            return m

    return None


def fetch(url):
    r = requests.get(url, headers=HEADERS, timeout=30, verify = False)
    r.raise_for_status()
    return r.text


def discover_series_urls(page):
    html = fetch(START_URL.format(page=page))
    soup = BeautifulSoup(html, "html.parser")

    urls = []
    for a in soup.find_all("a", href=True):
        href = a["href"]

        if re.search(r"/series/[0-9a-z]{7}", href):
            full = urljoin(BASE, href).split("?")[0]
            urls.append(full)

    return list(dict.fromkeys(urls))


def parse_series(url):
    html = fetch(url)
    soup = BeautifulSoup(html, "html.parser")

    title = None
    h1 = soup.find("h1")
    if h1:
        title = clean(h1.get_text(" ", strip=True))
    if not title and soup.title:
        title = clean(soup.title.get_text())

    raw_status = text_after_label(soup, "Status in Country of Origin")
    status_data = parse_status(raw_status)

    data = {
        "url": url,
        "title": clean_title(title),
        "type": text_after_label(soup, "Type"),
        "year": text_after_label(soup, "Year"),
        "volumes": status_data["volumes"],
        "chapters": status_data["chapters"],
        "status": status_data["status"],
        "raw_status": status_data["raw"],
        "authors": links_after_label(soup, "Author(s)", "/author/"),
        "artists": links_after_label(soup, "Artist(s)", "/author/"),
        "genres": links_after_label(
            soup,
            "Genre",
            "genre=",
            exclude_text=["Search for series"],
        ),
        "categories": links_after_label(soup, "Categories", "category="),
        "original_publisher": links_after_label(soup, "Original Publisher"),
        "english_publisher": links_after_label(soup, "English Publisher"),
        "description": text_after_label(soup, "Description"),
        "cover_image": find_cover(soup),
    }

    return data


def init_xlsx():
    if Path(OUT_XLSX).exists():
        return load_workbook(OUT_XLSX)

    wb = Workbook()
    ws = wb.active
    ws.title = "MangaUpdates"

    ws.append([
        "url",
        "title",
        "type",
        "year",
        "volumes",
        "chapters",
        "status",
        "raw_status",
        "authors",
        "artists",
        "genres",
        "categories",
        "original_publisher",
        "english_publisher",
        "description",
        "cover_image",
    ])

    wb.save(OUT_XLSX)
    return wb


def append_row(wb, data):
    ws = wb["MangaUpdates"]

    ws.append([
        data["url"],
        data["title"],
        data["type"],
        data["year"],
        data["volumes"],
        data["chapters"],
        data["status"],
        data["raw_status"],
        ", ".join(data["authors"]),
        ", ".join(data["artists"]),
        ", ".join(data["genres"]),
        ", ".join(data["categories"]),
        ", ".join(data["original_publisher"]),
        ", ".join(data["english_publisher"]),
        data["description"],
        data["cover_image"],
    ])


def load_seen():
    if Path(SEEN_FILE).exists():
        return set(json.loads(Path(SEEN_FILE).read_text()))
    return set()


def save_seen(seen):
    Path(SEEN_FILE).write_text(json.dumps(sorted(seen), indent=2))


def existing_count(wb):
    ws = wb["MangaUpdates"]
    return max(ws.max_row - 1, 0)


def polite_sleep():
    time.sleep(random.uniform(1.5, 3.5))


def main():
    wb = init_xlsx()
    seen = load_seen()
    saved = existing_count(wb)

    page = 1

    with tqdm(total=TARGET, initial=saved) as bar:
        while saved < TARGET:
            try:
                series_urls = discover_series_urls(page)
            except Exception as e:
                print(f"Failed listing page {page}: {e}")
                page += 1
                polite_sleep()
                continue

            if not series_urls:
                print(f"No series URLs found on page {page}. Stopping.")
                break

            for url in series_urls:
                if saved >= TARGET:
                    break

                if url in seen:
                    continue

                try:
                    data = parse_series(url)
                    seen.add(url)

                    if clean(data.get("type")).lower() != "manga":
                        save_seen(seen)
                        polite_sleep()
                        continue

                    append_row(wb, data)
                    saved += 1
                    bar.update(1)

                    if saved % 25 == 0:
                        wb.save(OUT_XLSX)
                        save_seen(seen)

                except Exception as e:
                    print(f"Failed {url}: {e}")

                polite_sleep()

            page += 1
            wb.save(OUT_XLSX)
            save_seen(seen)

    wb.save(OUT_XLSX)
    save_seen(seen)
    print(f"Saved {saved} manga rows to {OUT_XLSX}")


if __name__ == "__main__":
    main()